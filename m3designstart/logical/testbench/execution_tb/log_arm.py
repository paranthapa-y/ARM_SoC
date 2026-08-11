#!/usr/bin/env python3

import re
import sys
import openpyxl


# =========================================================
# Configuration
# =========================================================

if len(sys.argv) != 2:
    print("Usage: python3 log_arm.py <TESTNAME>")
    sys.exit(1)

TESTNAME = sys.argv[1]

TRACKER_LOG = "processor_log.txt"
DISASM_FILE = f"{TESTNAME}.dis"
MEMORY_MAP_FILE = "memory_map.xlsx"
OUTPUT_FILE = f"{TESTNAME}_log_decoded.txt"


# =========================================================
# Load Memory Map
# =========================================================

print(f"Loading memory map from {MEMORY_MAP_FILE}...")

try:
    workbook = openpyxl.load_workbook(
        MEMORY_MAP_FILE,
        data_only=True
    )

    sheet = workbook.active

except Exception as e:
    print(f"ERROR: Cannot open memory map: {e}")
    sys.exit(1)


memory_map = []


# =========================================================
# Find Required Columns
# =========================================================

headers = {}

for cell in sheet[1]:

    if cell.value is not None:
        headers[str(cell.value).strip().lower()] = cell.column


required_columns = [
    "peripheral",
    "start",
    "end"
]

for column in required_columns:

    if column not in headers:

        print(
            f"ERROR: Memory map must contain "
            f"'{column}' column"
        )

        sys.exit(1)


peripheral_col = headers["peripheral"]
start_col = headers["start"]
end_col = headers["end"]


# =========================================================
# Read Memory Map
# =========================================================

for row in range(2, sheet.max_row + 1):

    peripheral = sheet.cell(
        row=row,
        column=peripheral_col
    ).value

    start = sheet.cell(
        row=row,
        column=start_col
    ).value

    end = sheet.cell(
        row=row,
        column=end_col
    ).value


    if peripheral is None or start is None or end is None:
        continue


    try:

        if isinstance(start, str):
            start_addr = int(
                start.strip(),
                16
            )
        else:
            start_addr = int(start)


        if isinstance(end, str):
            end_addr = int(
                end.strip(),
                16
            )
        else:
            end_addr = int(end)


    except ValueError:

        print(
            f"WARNING: Invalid address "
            f"in memory map row {row}"
        )

        continue


    memory_map.append(
        (
            start_addr,
            end_addr,
            str(peripheral).strip()
        )
    )


print(
    f"Loaded {len(memory_map)} "
    f"memory map entries"
)


# =========================================================
# Find Peripheral From Address
# =========================================================

def find_peripheral(address):

    for start_addr, end_addr, peripheral in memory_map:

        if start_addr <= address <= end_addr:

            return peripheral

    return "UNKNOWN"


# =========================================================
# Convert ARM Instruction to Uppercase
#
# Example:
#
# ldr r1, [pc, #44]
#
# becomes:
#
# LDR R1, [PC, #44]
#
# But:
#
# .word 0x2000fc00
#
# remains:
#
# .word 0x2000fc00
# =========================================================

def format_assembly(asm):

    asm = asm.strip()


    # -----------------------------------------------------
    # .word must remain unchanged
    # -----------------------------------------------------

    if asm.lower().startswith(".word"):

        return asm


    # -----------------------------------------------------
    # Split mnemonic and operands
    # -----------------------------------------------------

    parts = asm.split(" ", 1)


    # -----------------------------------------------------
    # Only mnemonic
    # -----------------------------------------------------

    if len(parts) == 1:

        return parts[0].upper()


    mnemonic = parts[0].upper()

    operands = parts[1]


    # -----------------------------------------------------
    # Convert ARM registers to uppercase
    #
    # r0-r15
    # pc
    # sp
    # lr
    # -----------------------------------------------------

    operands = re.sub(
        r'\b(r(?:1[0-5]|[0-9]))\b',
        lambda m: m.group(1).upper(),
        operands,
        flags=re.IGNORECASE
    )


    operands = re.sub(
        r'\b(pc|sp|lr)\b',
        lambda m: m.group(1).upper(),
        operands,
        flags=re.IGNORECASE
    )


    return mnemonic + " " + operands


# =========================================================
# Build Instruction Map From objdump
# =========================================================

instruction_map = {}


pattern = re.compile(
    r'^\s*'
    r'([0-9a-fA-F]+):'
    r'\s+'
    r'([0-9a-fA-F ]+)'
    r'\s+'
    r'(.+)$'
)


print(
    f"Loading disassembly from {DISASM_FILE}..."
)


try:

    with open(DISASM_FILE) as f:

        for line in f:

            match = pattern.match(line)

            if not match:
                continue


            # -------------------------------------------------
            # Address
            # -------------------------------------------------

            addr = int(
                match.group(1),
                16
            )


            # -------------------------------------------------
            # Opcode
            # -------------------------------------------------

            opcode = match.group(2).strip()


            # -------------------------------------------------
            # Assembly
            # -------------------------------------------------

            asm = match.group(3).strip()


            # -------------------------------------------------
            # Remove comments
            # -------------------------------------------------

            asm = asm.split(";")[0].strip()


            # -------------------------------------------------
            # Remove symbol names
            #
            # Example:
            #
            # bl 2f0 <Uart1Putc>
            #
            # becomes:
            #
            # bl 2f0
            # -------------------------------------------------

            asm = re.sub(
                r'<[^>]+>',
                '',
                asm
            )


            # -------------------------------------------------
            # Compress multiple spaces
            # -------------------------------------------------

            asm = re.sub(
                r'\s+',
                ' ',
                asm
            ).strip()


            # -------------------------------------------------
            # Format ARM instruction
            # -------------------------------------------------

            asm = format_assembly(asm)


            # -------------------------------------------------
            # Determine instruction size
            # -------------------------------------------------

            words = opcode.split()


            # 16-bit Thumb instruction
            if len(words) == 1:

                size = 2


            # 32-bit Thumb-2 instruction
            elif len(words) == 2:

                size = 4


            else:

                continue


            instruction_map[addr] = (
                size,
                opcode,
                asm
            )


except FileNotFoundError:

    print(
        f"ERROR: Disassembly file "
        f"'{DISASM_FILE}' not found"
    )

    sys.exit(1)


print(
    f"Loaded {len(instruction_map)} "
    f"instructions"
)


# =========================================================
# Open Tracker Log
# =========================================================

print(
    f"Reading tracker log: {TRACKER_LOG}"
)


try:

    fin = open(
        TRACKER_LOG,
        "r"
    )

except FileNotFoundError:

    print(
        f"ERROR: Tracker log "
        f"'{TRACKER_LOG}' not found"
    )

    sys.exit(1)


fout = open(
    OUTPUT_FILE,
    "w"
)


# =========================================================
# Output Header
# =========================================================

fout.write(
    "-" * 150 + "\n"
)


fout.write(
    f"{'TIME':<15}"
    f"{'BUS':<6}"
    f"{'TYPE':<8}"
    f"{'ADDRESS':<14}"
    f"{'DATA':<14}"
    f"{'OPCODE':<14}"
    f"{'INSTRUCTION':<35}"
    f"{'PERIPHERAL'}\n"
)


fout.write(
    "-" * 150 + "\n"
)


# =========================================================
# Process Tracker Log
# =========================================================

for line in fin:


    # -----------------------------------------------------
    # Ignore original header/separators
    # -----------------------------------------------------

    if line.startswith("-"):
        continue


    if line.startswith("TIME"):
        continue


    if line.strip() == "":
        continue


    cols = line.split()


    # -----------------------------------------------------
    # Expected:
    #
    # TIME BUS TYPE ADDRESS DATA INSTRUCTION RESP
    # -----------------------------------------------------

    if len(cols) < 7:
        continue


    time = cols[0]

    bus = cols[1]

    rw = cols[2]

    addr = cols[3]

    data = cols[4]


    # -----------------------------------------------------
    # Convert address to integer
    # -----------------------------------------------------

    try:

        address = int(
            addr,
            16
        )

    except ValueError:

        continue


    # =====================================================
    # D-BUS / S-BUS
    # =====================================================

    if bus != "I":


        peripheral = find_peripheral(
            address
        )


        fout.write(
            f"{time:<15}"
            f"{bus:<6}"
            f"{rw:<8}"
            f"{addr:<14}"
            f"{data:<14}"
            f"{'-':<14}"
            f"{'-':<35}"
            f"{peripheral}\n"
        )


        continue


    # =====================================================
    # I-BUS
    # =====================================================

    fetch_addr = address


    # =====================================================
    # Vector Table
    #
    # These are data words, not instructions.
    # =====================================================

    if fetch_addr < 0x124:


        peripheral = find_peripheral(
            fetch_addr
        )


        fout.write(
            f"{time:<15}"
            f"{bus:<6}"
            f"{rw:<8}"
            f"{addr:<14}"
            f"{data:<14}"
            f"{cols[5]:<14}"
            f"{('.word ' + cols[5]):<35}"
            f"{peripheral}\n"
        )


        continue


    # =====================================================
    # Decode Instructions
    # =====================================================

    pc = fetch_addr


    while pc < fetch_addr + 4:


        # -------------------------------------------------
        # No instruction at this address
        # -------------------------------------------------

        if pc not in instruction_map:

            pc += 2

            continue


        # -------------------------------------------------
        # Get instruction
        # -------------------------------------------------

        size, opcode, asm = instruction_map[pc]


        # -------------------------------------------------
        # Determine peripheral
        # -------------------------------------------------

        peripheral = find_peripheral(
            pc
        )


        # -------------------------------------------------
        # Write decoded instruction
        # -------------------------------------------------

        fout.write(
            f"{time:<15}"
            f"{bus:<6}"
            f"{rw:<8}"
            f"{f'0x{pc:08x}':<14}"
            f"{data:<14}"
            f"{opcode:<14}"
            f"{asm:<35}"
            f"{peripheral}\n"
        )


        # -------------------------------------------------
        # Move to next instruction
        #
        # 16-bit Thumb  -> +2
        # 32-bit Thumb2 -> +4
        # -------------------------------------------------

        pc += size


# =========================================================
# Close Files
# =========================================================

fin.close()

fout.close()


# =========================================================
# Done
# =========================================================

print()
print(
    "=============================================="
)

print(
    "Tracker decoding completed"
)

print(
    "=============================================="
)

print(
    f"Output file : {OUTPUT_FILE}"
)

print(
    f"Memory map  : {MEMORY_MAP_FILE}"
)

print(
    f"Disassembly : {DISASM_FILE}"
)

print(
    "=============================================="
)
